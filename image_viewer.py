import tkinter as tk
from tkinter import ttk
import os
import threading
import time
from datetime import datetime, timedelta
import tempfile
import shutil
import subprocess
import logging
from logging.handlers import RotatingFileHandler
import gc
import re
import hashlib
from PIL import Image, ImageTk

LOG_FILE = "/var/log/pi-photo-viewer/app.log"
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

try:
    handler = RotatingFileHandler(LOG_FILE, maxBytes=5*1024*1024, backupCount=5)
    formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
    handler.setFormatter(formatter)
    logger.addHandler(handler)
except Exception as e:
    logging.basicConfig(level=logging.INFO)
    logger.warning(f"Could not create file logger: {e}")

NETWORK_BASE_PATH = "/mnt/network-drive/TS16949 Work/Standard Operating Procedure"
DEPARTMENTS = ["11 Injection", "12 Assembly", "13 Paint", "32 Repack", "New Model", "SHOOT AND SHIP"]

# Special department paths (for departments nested within other department folders)
# Format: "Department Name": "relative/path/from/base"
# Example: "SHOOT AND SHIP" appears as its own department but lives in 11 Injection folder
SPECIAL_DEPT_PATHS = {
    "SHOOT AND SHIP": "11 Injection/SHOOT AND SHIP"
}

# Folders to exclude when listing models (prevents nested departments from appearing as models)
# Format: "Parent Department": ["folder1", "folder2"]
# When "11 Injection" is selected, "SHOOT AND SHIP" won't appear in the model list
EXCLUDED_MODEL_FOLDERS = {
    "11 Injection": ["SHOOT AND SHIP"]
}

SUPPORTED_FORMATS = (".xlsx", ".png", ".jpg", ".jpeg", ".gif", ".bmp")
LOGO_WIDTH = 175
IMAGE_CACHE_SIZE = 20  # Increased from 2 to 20 - keeps 10 files (front+back) in memory
MAX_IMAGE_DIMENSION = 1920
CACHE_STALE_DAYS = 7

SHEET_MAPPING = {
    "front": ["front", "front page", "proposal"],
    "back": ["back", "back page"],
}

class ImageCache:
    def __init__(self, max_size=IMAGE_CACHE_SIZE):
        self.cache = {}
        self.max_size = max_size
        self.order = []
        logger.info(f"Memory cache initialized (max size: {max_size} images)")
    
    def get(self, path):
        if path in self.cache:
            self.order.remove(path)
            self.order.append(path)
            logger.info(f"[INSTANT LOAD] Memory cache hit - displaying immediately!")
            return self.cache[path]
        return None
    
    def put(self, path, photo):
        if path in self.cache:
            self.order.remove(path)
        elif len(self.cache) >= self.max_size:
            removed = self.order.pop(0)
            if removed in self.cache:
                logger.debug(f"Evicting from cache: {removed}")
                del self.cache[removed]
        
        self.cache[path] = photo
        self.order.append(path)
        logger.debug(f"Memory cache now contains {len(self.cache)}/{self.max_size} images")

class ExcelConverter:
    def __init__(self, cache_dir="/tmp/pi-photo-viewer-cache"):
        self.cache_dir = cache_dir
        self.conversion_lock = threading.Lock()
        os.makedirs(cache_dir, exist_ok=True)
        self._check_tools()
        self._log_cache_status()
    
    def _log_cache_status(self):
        """Log cache directory status on startup"""
        try:
            files = os.listdir(self.cache_dir)
            png_files = [f for f in files if f.endswith('.png')]
            meta_files = [f for f in files if f.endswith('.meta')]
            logger.info(f"Cache directory: {self.cache_dir}")
            logger.info(f"Cache contains: {len(png_files)} PNG files, {len(meta_files)} metadata files")
        except Exception as e:
            logger.error(f"Error checking cache status: {e}")
    
    def _check_tools(self):
        """Verify required tools are available - checks both 'convert' and 'magick'"""
        tools_required = ["libreoffice", "pdftoppm"]
        tools_optional = ["convert", "magick"]
        
        missing = []
        for tool in tools_required:
            if not shutil.which(tool):
                missing.append(tool)
        
        # Check for ImageMagick (either convert or magick)
        self.imagemagick_cmd = None
        if shutil.which("convert"):
            self.imagemagick_cmd = "convert"
        elif shutil.which("magick"):
            self.imagemagick_cmd = "magick"
        else:
            missing.append("convert/magick")
        
        if missing:
            logger.error(f"Missing required tools: {', '.join(missing)}")
    
    def sanitize_filename(self, name):
        """Remove unsafe characters and add hash for uniqueness"""
        safe = re.sub(r'[<>:"/\\|?*]', '_', name)[:80]
        return safe
    
    def find_sheet(self, excel_path, sheet_type):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            
            patterns = SHEET_MAPPING.get(sheet_type.lower(), [])
            
            for pattern in patterns:
                for sheet_name in sheet_names:
                    if pattern.lower() in sheet_name.lower():
                        logger.debug(f"Matched '{sheet_type}' to '{sheet_name}'")
                        return sheet_name
            
            logger.warning(f"No sheet match for '{sheet_type}' in {os.path.basename(excel_path)}")
            return None
        except Exception as e:
            logger.error(f"Error reading sheets from {os.path.basename(excel_path)}: {e}")
            return None
    
    def get_cache_path(self, excel_path, sheet_name):
        """Generate safe cache filename with hash"""
        safe_name = self.sanitize_filename(f"{os.path.basename(excel_path)}_{sheet_name}")
        content_hash = hashlib.sha1(f"{excel_path}_{sheet_name}".encode()).hexdigest()[:8]
        return os.path.join(self.cache_dir, f"{safe_name}_{content_hash}.png")
    
    def get_meta_path(self, cache_png_path):
        """Get metadata file path"""
        base = os.path.splitext(cache_png_path)[0]
        return base + ".meta"
    
    def get_sheet_index(self, excel_path, sheet_name):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            return sheet_names.index(sheet_name) if sheet_name in sheet_names else None
        except Exception as e:
            logger.error(f"Error getting sheet index: {e}")
            return None
    
    def is_cache_valid(self, cache_path, source_excel_path=None):
        """Check if cache is valid with detailed logging"""
        cache_name = os.path.basename(cache_path)
        
        # Check if cache file exists
        if not os.path.exists(cache_path):
            return False
        
        # Check source modification time if provided
        if source_excel_path and os.path.exists(source_excel_path):
            try:
                excel_mod_time = os.path.getmtime(source_excel_path)
                meta_path = self.get_meta_path(cache_path)
                
                if os.path.exists(meta_path):
                    try:
                        with open(meta_path, 'r') as f:
                            cached_time = float(f.read().strip())
                        
                        if excel_mod_time > cached_time:
                            logger.info(f"Cache invalid: {cache_name} - source modified")
                            return False
                    except Exception as e:
                        logger.debug(f"Metadata read error: {e}")
            except Exception as e:
                logger.debug(f"Error checking source time: {e}")
        
        # Check file age
        try:
            file_age = datetime.now() - datetime.fromtimestamp(os.path.getmtime(cache_path))
            if file_age >= timedelta(days=CACHE_STALE_DAYS):
                logger.info(f"Cache expired: {cache_name} - {file_age.days} days old")
                return False
            
            logger.debug(f"[CACHE HIT] {cache_name}")
            return True
        except Exception as e:
            logger.error(f"Cache validation error: {e}")
            return False
    
    def save_metadata(self, cache_path, source_excel_path):
        """Save source modification time"""
        try:
            excel_mod_time = os.path.getmtime(source_excel_path)
            meta_path = self.get_meta_path(cache_path)
            with open(meta_path, 'w') as f:
                f.write(str(excel_mod_time))
        except Exception as e:
            logger.error(f"Error saving metadata: {e}")
    
    def convert_excel_to_png(self, excel_path, sheet_name, stop_event=None):
        """Convert Excel sheet to PNG with optional cancellation support"""
        with self.conversion_lock:
            # Check if we should stop before starting
            if stop_event and stop_event.is_set():
                logger.debug("Conversion cancelled before start")
                return None
            
            temp_dir = None
            try:
                cache_path = self.get_cache_path(excel_path, sheet_name)
                
                if self.is_cache_valid(cache_path, excel_path):
                    # Cache hit - return immediately (file exists on disk)
                    return cache_path
                
                # Cache miss - need to convert
                logger.info(f"[CONVERTING] {os.path.basename(excel_path)} - {sheet_name}")
                sheet_index = self.get_sheet_index(excel_path, sheet_name)
                if sheet_index is None:
                    logger.error(f"Sheet '{sheet_name}' not found in workbook")
                    return None
                
                # Check stop event before expensive operations
                if stop_event and stop_event.is_set():
                    logger.info("Conversion cancelled during processing")
                    return None
                
                temp_dir = tempfile.mkdtemp()
                output_prefix = os.path.splitext(cache_path)[0]
                
                cmd = ["libreoffice", "--headless", "--invisible", "--nocrashreport",
                       "--nodefault", "--nofirststartwizard", "--nologo", "--norestore",
                       "--convert-to", "pdf", "--outdir", temp_dir, excel_path]
                result = subprocess.run(cmd, capture_output=True, timeout=45, text=True)
                
                if result.returncode != 0:
                    logger.error(f"LibreOffice failed: {result.stderr[:400]}")
                    return None
                
                # Check stop event after LibreOffice
                if stop_event and stop_event.is_set():
                    logger.info("Conversion cancelled after PDF generation")
                    return None
                
                pdf_files = [f for f in os.listdir(temp_dir) if f.endswith(".pdf")]
                if not pdf_files:
                    logger.error("No PDF generated")
                    return None
                
                pdf_path = os.path.join(temp_dir, pdf_files[0])
                pdf_page = sheet_index + 1
                
                cmd = ["pdftoppm", "-png", "-f", str(pdf_page), "-l", str(pdf_page),
                       "-singlefile", "-r", "150", pdf_path, output_prefix]
                result = subprocess.run(cmd, capture_output=True, timeout=30, text=True)
                
                if result.returncode != 0 and self.imagemagick_cmd:
                    logger.warning(f"pdftoppm failed: {result.stderr[:400]}, trying ImageMagick")
                    cmd = [self.imagemagick_cmd, "-density", "100", f"{pdf_path}[{sheet_index}]", cache_path]
                    result = subprocess.run(cmd, capture_output=True, timeout=30, text=True)
                    if result.returncode != 0:
                        logger.error(f"ImageMagick failed: {result.stderr[:400]}")
                        return None
                
                if os.path.exists(cache_path):
                    self.save_metadata(cache_path, excel_path)
                    logger.info(f"[CACHED] {os.path.basename(cache_path)}")
                    return cache_path
                else:
                    logger.error(f"Conversion completed but file not found: {cache_path}")
                
                return None
            
            except subprocess.TimeoutExpired:
                logger.error("Conversion timeout")
                return None
            except Exception as e:
                logger.error(f"Conversion error: {e}")
                return None
            finally:
                if temp_dir and os.path.exists(temp_dir):
                    try:
                        shutil.rmtree(temp_dir, ignore_errors=True)
                    except Exception:
                        pass

class FullscreenImageApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Pi Standards Viewer - Network")
        logger.info("App starting")
        
        self.root.after(200, lambda: self.root.attributes("-fullscreen", True))
        self.root.configure(bg="black")
        
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TCombobox", fieldbackground="#374151", background="#4B5563",
                       foreground="#06B6D4", arrowcolor="#06B6D4", arrowsize=30)
        self.root.option_add("*TCombobox*Listbox.font", ("Helvetica", 22, "bold"))
        self.root.option_add("*TCombobox*Listbox.background", "#1F2937")
        self.root.option_add("*TCombobox*Listbox.foreground", "#06B6D4")
        
        self.control_bar_collapsed_height = 80
        self.control_bar_expanded_height = 220
        self.collapse_timer = None
        self.collapse_delay = 30000
        
        self.control_frame = tk.Frame(root, bg="#1F2937", pady=10, padx=30)
        self.control_frame.pack(side="bottom", fill="x")
        self.control_frame.pack_propagate(False)
        self.control_frame.config(height=self.control_bar_collapsed_height)
        
        self.expand_indicator = tk.Label(self.control_frame, text="▲ TAP TO SELECT FILE ▲",
                                         bg="#1F2937", fg="#06B6D4", font=("Helvetica", 14, "bold"),
                                         cursor="hand2")
        self.expand_indicator.pack(side="top", pady=(0, 5))
        self.expand_indicator.bind("<Button-1>", lambda e: self.expand_controls())
        
        self.collapsed_container = tk.Frame(self.control_frame, bg="#1F2937")
        self.collapsed_container.pack(fill="both", expand=True)
        
        self.expanded_container = tk.Frame(self.control_frame, bg="#1F2937")
        
        collapsed_label = tk.Label(self.collapsed_container, text="Page:", bg="#1F2937",
                                  fg="#06B6D4", font=("Helvetica", 20, "bold"))
        collapsed_label.pack(side="left", padx=(0, 15))
        
        self.collapsed_button_frame = tk.Frame(self.collapsed_container, bg="#1F2937")
        self.collapsed_button_frame.pack(side="left", expand=True, fill="both")
        
        self.front_button = tk.Button(self.collapsed_button_frame, text="FRONT PAGE",
                                     font=("Helvetica", 22, "bold"), bg="#06B6D4", fg="#1F2937",
                                     activebackground="#0891B2", relief="sunken", bd=3, cursor="hand2",
                                     command=lambda: self.on_page_click("Front"))
        self.back_button = tk.Button(self.collapsed_button_frame, text="BACK PAGE",
                                    font=("Helvetica", 22, "bold"), bg="#4B5563", fg="#06B6D4",
                                    activebackground="#374151", relief="raised", bd=3, cursor="hand2",
                                    command=lambda: self.on_page_click("Back"))
        
        self.front_button.pack(side="left", expand=True, fill="both", padx=(0, 10))
        self.back_button.pack(side="left", expand=True, fill="both")
        
        row1 = tk.Frame(self.expanded_container, bg="#1F2937")
        row1.pack(fill="x", pady=(0, 10))
        
        tk.Label(row1, text="Department:", bg="#1F2937", fg="#06B6D4",
                font=("Helvetica", 20, "bold")).pack(side="left", padx=(0, 10))
        self.dept_var = tk.StringVar(root)
        self.dept_dropdown = ttk.Combobox(row1, textvariable=self.dept_var,
                                         font=("Helvetica", 22, "bold"), state="readonly", height=5,
                                         postcommand=self.on_dropdown_open)
        self.dept_dropdown["values"] = DEPARTMENTS
        self.dept_dropdown.pack(side="left", expand=True, fill="both", padx=(0, 20))
        self.dept_dropdown.bind("<<ComboboxSelected>>", self.on_dept_select)
        
        tk.Label(row1, text="Part Model:", bg="#1F2937", fg="#06B6D4",
                font=("Helvetica", 20, "bold")).pack(side="left", padx=(0, 10))
        self.model_var = tk.StringVar(root)
        self.model_dropdown = ttk.Combobox(row1, textvariable=self.model_var,
                                          font=("Helvetica", 22, "bold"), state="readonly", height=5,
                                          postcommand=self.on_dropdown_open)
        self.model_dropdown.pack(side="left", expand=True, fill="both")
        self.model_dropdown.bind("<<ComboboxSelected>>", self.on_model_select)
        
        row2 = tk.Frame(self.expanded_container, bg="#1F2937")
        row2.pack(fill="x", pady=(0, 10))
        
        tk.Label(row2, text="File:", bg="#1F2937", fg="#06B6D4",
                font=("Helvetica", 20, "bold")).pack(side="left", padx=(0, 10))
        self.file_var = tk.StringVar(root)
        self.file_dropdown = ttk.Combobox(row2, textvariable=self.file_var,
                                         font=("Helvetica", 22, "bold"), state="readonly", height=5,
                                         postcommand=self.on_dropdown_open)
        self.file_dropdown.pack(side="left", expand=True, fill="both")
        self.file_dropdown.bind("<<ComboboxSelected>>", self.on_file_select)
        
        row3 = tk.Frame(self.expanded_container, bg="#1F2937")
        row3.pack(fill="x")
        tk.Label(row3, text="Page:", bg="#1F2937", fg="#06B6D4",
                font=("Helvetica", 20, "bold")).pack(side="left", padx=(0, 10))
        
        self.button_frame = tk.Frame(row3, bg="#1F2937")
        self.button_frame.pack(side="left", expand=True, fill="both")
        
        self.front_button_exp = tk.Button(self.button_frame, text="FRONT PAGE",
                                         font=("Helvetica", 22, "bold"), bg="#06B6D4", fg="#1F2937",
                                         activebackground="#0891B2", relief="sunken", bd=3, cursor="hand2",
                                         command=lambda: self.on_page_click("Front"))
        self.back_button_exp = tk.Button(self.button_frame, text="BACK PAGE",
                                        font=("Helvetica", 22, "bold"), bg="#4B5563", fg="#06B6D4",
                                        activebackground="#374151", relief="raised", bd=3, cursor="hand2",
                                        command=lambda: self.on_page_click("Back"))
        
        self.front_button_exp.pack(side="left", expand=True, fill="both", padx=(0, 10))
        self.back_button_exp.pack(side="left", expand=True, fill="both")
        
        self.image_label = tk.Label(root, bg="black", fg="white", font=("Helvetica", 24))
        self.image_label.pack(expand=True, fill="both")
        self.image_label.bind("<Button-1>", lambda e: self.expand_controls())
        self.image_label.config(image="", text="Waiting for network drive...\n(polling every 10 seconds)")

        self.current_page = "Front"
        self.is_expanded = False
        self.current_file_path = None
        self.current_model_path = None
        self.files_list = []
        self.image_cache = ImageCache()
        self.excel_converter = ExcelConverter()

        # Thread management with explicit per-thread stop events
        self.bg_precache_thread = None
        self.bg_precache_stop = None
        self.fg_precache_thread = None
        self.fg_precache_stop = None
        self.polling_thread = None
        self.polling_stop = threading.Event()

        self.network_available = False
        self.set_online_state(False)

        if DEPARTMENTS:
            self.dept_var.set(DEPARTMENTS[0])
            self.root.after_idle(self.start_network_polling)
        
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)
    
    def on_dropdown_open(self):
        if self.collapse_timer:
            self.root.after_cancel(self.collapse_timer)
        self.reset_collapse_timer()
    
    def expand_controls(self):
        if self.is_expanded:
            return
        
        # Close any open dropdown menus before expanding
        self.root.focus_set()
        try:
            self.dept_dropdown.selection_clear()
            self.model_dropdown.selection_clear()
            self.file_dropdown.selection_clear()
        except:
            pass
        
        self.is_expanded = True
        self.collapsed_container.pack_forget()
        self.expanded_container.pack(fill="both", expand=True)
        self.expand_indicator.config(text="▼ TAP IMAGE TO HIDE ▼")
        self.control_frame.config(height=self.control_bar_expanded_height)
        self.reset_collapse_timer()
    
    def collapse_controls(self):
        if not self.is_expanded:
            return
        
        # Close any open dropdown menus by shifting focus
        self.root.focus_set()
        
        # Force close dropdown popups
        try:
            self.dept_dropdown.selection_clear()
            self.model_dropdown.selection_clear()
            self.file_dropdown.selection_clear()
        except:
            pass
        
        self.is_expanded = False
        self.expanded_container.pack_forget()
        self.collapsed_container.pack(fill="both", expand=True)
        self.expand_indicator.config(text="▲ TAP TO SELECT FILE ▲")
        self.control_frame.config(height=self.control_bar_collapsed_height)
        if self.collapse_timer:
            self.root.after_cancel(self.collapse_timer)
            self.collapse_timer = None
    
    def reset_collapse_timer(self):
        if self.collapse_timer:
            self.root.after_cancel(self.collapse_timer)
        self.collapse_timer = self.root.after(self.collapse_delay, self.collapse_controls)
    
    def on_page_click(self, page):
        self.current_page = page
        for front, back in [(self.front_button, self.back_button),
                           (self.front_button_exp, self.back_button_exp)]:
            if page == "Front":
                front.config(bg="#06B6D4", fg="#1F2937", relief="sunken")
                back.config(bg="#4B5563", fg="#06B6D4", relief="raised")
            else:
                back.config(bg="#06B6D4", fg="#1F2937", relief="sunken")
                front.config(bg="#4B5563", fg="#06B6D4", relief="raised")
        
        if self.current_file_path:
            self.display_file(self.current_file_path, page)
        if self.is_expanded:
            self.reset_collapse_timer()
    
    def on_close(self):
        logger.info("App closing - setting stop flags")

        # Stop polling
        self.polling_stop.set()
        
        # Stop background precache
        if self.bg_precache_stop:
            self.bg_precache_stop.set()
        if self.bg_precache_thread and self.bg_precache_thread.is_alive():
            self.bg_precache_thread.join(timeout=2.0)
        
        # Stop foreground precache
        if self.fg_precache_stop:
            self.fg_precache_stop.set()
        if self.fg_precache_thread and self.fg_precache_thread.is_alive():
            self.fg_precache_thread.join(timeout=2.0)
        
        self.root.quit()
    
    def set_online_state(self, online: bool):
        """Update UI controls based on network availability."""

        def apply_state():
            self.network_available = online
            state = "readonly" if online else "disabled"
            for combo in (self.dept_dropdown, self.model_dropdown, self.file_dropdown):
                combo.configure(state=state)

            if not online:
                self.image_label.config(image="", text="Waiting for network drive...\n(polling every 10 seconds)")

        if threading.current_thread() is threading.main_thread():
            apply_state()
        else:
            self.root.after(0, apply_state)

    def start_network_polling(self):
        """Start polling for network drive availability"""
        logger.info("Starting network drive polling")
        self.polling_thread = threading.Thread(target=self._poll_network_drive, daemon=True)
        self.polling_thread.start()
    
    def _poll_network_drive(self):
        """Poll network drive every 10 seconds until available"""
        while not self.polling_stop.is_set():
            try:
                if os.path.exists(NETWORK_BASE_PATH) and os.path.isdir(NETWORK_BASE_PATH):
                    logger.info("Network drive found!")
                    self.root.after(0, lambda: self.set_online_state(True))
                    self.root.after(0, lambda: self.on_dept_select(None))
                    return
                else:
                    logger.debug("Network drive not available yet")
                    self.root.after(0, lambda: self.set_online_state(False))
            except Exception as e:
                logger.debug(f"Network poll error: {e}")
                self.root.after(0, lambda: self.set_online_state(False))
            
            for _ in range(100):
                if self.polling_stop.is_set():
                    return
                time.sleep(0.1)
    
    def on_dept_select(self, event):
        logger.info(f"Department: {self.dept_var.get()}")
        threading.Thread(target=self._dept_select_worker, daemon=True).start()
    
    def _dept_select_worker(self):
        """Background worker for department selection"""
        try:
            self.update_models()
        except Exception as e:
            logger.error(f"Error updating models: {e}")
            self.root.after(0, lambda: self.image_label.config(image="", text="Network drive not available"))
            return
        
        # Stop old background precache thread safely
        if self.bg_precache_stop and self.bg_precache_thread:
            logger.debug("Stopping old bg precache thread")
            self.bg_precache_stop.set()
            self.bg_precache_thread.join(timeout=2.0)
        
        # Start new background precache with fresh stop event
        dept = self.dept_var.get()
        if dept:
            logger.info(f"Starting bg precache for {dept}")
            self.bg_precache_stop = threading.Event()
            self.bg_precache_thread = threading.Thread(
                target=self.precache_dept,
                args=(dept, self.bg_precache_stop),
                daemon=True
            )
            self.bg_precache_thread.start()
        
        if self.is_expanded:
            self.root.after(0, self.reset_collapse_timer)
    
    def update_models(self):
        dept = self.dept_var.get()
        if not dept:
            self.root.after(0, lambda: self.model_dropdown.__setitem__("values", []))
            return
        
        # Check if this is a special department with a custom path
        if dept in SPECIAL_DEPT_PATHS:
            dept_path = os.path.join(NETWORK_BASE_PATH, SPECIAL_DEPT_PATHS[dept])
        else:
            dept_path = os.path.join(NETWORK_BASE_PATH, dept)
        
        if not os.path.exists(dept_path):
            logger.error(f"Department path not found: {dept_path}")
            self.root.after(0, lambda: self.model_dropdown.__setitem__("values", []))
            self.root.after(0, lambda: self.image_label.config(image="", text="Department not accessible"))
            return
        
        try:
            # Get all subdirectories
            all_models = [d for d in os.listdir(dept_path)
                         if os.path.isdir(os.path.join(dept_path, d))]
            
            # Filter out excluded folders for this department
            excluded = EXCLUDED_MODEL_FOLDERS.get(dept, [])
            models = sorted([m for m in all_models if m not in excluded])
            
            self.root.after(0, lambda: self.model_dropdown.__setitem__("values", models))
            if models:
                self.root.after(0, lambda: (self.model_var.set(models[0]), self.on_model_select(None)))
        except Exception as e:
            logger.error(f"Error listing models: {e}")
            self.root.after(0, lambda: self.model_dropdown.__setitem__("values", []))
            self.root.after(0, lambda: self.image_label.config(image="", text="Error reading department"))
    
    def on_model_select(self, event):
        # Stop old foreground precache
        if self.fg_precache_stop and self.fg_precache_thread:
            logger.debug("Stopping old fg precache thread")
            self.fg_precache_stop.set()
            self.fg_precache_thread.join(timeout=1.0)
        
        self.update_files()
        
        # Start new foreground precache
        if self.current_model_path:
            logger.info(f"Starting fg precache for {os.path.basename(self.current_model_path)}")
            self.fg_precache_stop = threading.Event()
            self.fg_precache_thread = threading.Thread(
                target=self.precache_model_aggressive,
                args=(self.current_model_path, self.fg_precache_stop),
                daemon=True
            )
            self.fg_precache_thread.start()
        
        if self.is_expanded:
            self.root.after(0, self.reset_collapse_timer)
    
    def update_files(self):
        dept = self.dept_var.get()
        model = self.model_var.get()
        if not dept or not model:
            self.file_dropdown["values"] = []
            self.files_list = []
            return
        
        # Check if this is a special department with a custom path
        if dept in SPECIAL_DEPT_PATHS:
            dept_path = os.path.join(NETWORK_BASE_PATH, SPECIAL_DEPT_PATHS[dept])
        else:
            dept_path = os.path.join(NETWORK_BASE_PATH, dept)
        
        self.current_model_path = os.path.join(dept_path, model)
        
        if not os.path.exists(self.current_model_path):
            logger.error(f"Model path not found: {self.current_model_path}")
            self.file_dropdown["values"] = []
            self.files_list = []
            self.root.after(0, lambda: self.image_label.config(image="", text="Model not accessible"))
            return
        
        try:
            self.files_list = sorted([os.path.join(self.current_model_path, f)
                                     for f in os.listdir(self.current_model_path)
                                     if os.path.isfile(os.path.join(self.current_model_path, f))
                                     and f.lower().endswith(SUPPORTED_FORMATS)])
            names = [os.path.splitext(os.path.basename(f))[0] for f in self.files_list]
            self.file_dropdown["values"] = names
            if names:
                self.file_var.set(names[0])
                self.on_file_select(None)
            else:
                self.root.after(0, lambda: self.image_label.config(image="", text="No files found"))
        except Exception as e:
            logger.error(f"Error listing files: {e}")
            self.file_dropdown["values"] = []
            self.files_list = []
            self.root.after(0, lambda: self.image_label.config(image="", text="Error reading files"))
    
    def on_file_select(self, event):
        name = self.file_var.get()
        if not name or not self.files_list:
            return
        
        for f in self.files_list:
            if os.path.splitext(os.path.basename(f))[0] == name:
                self.current_file_path = f
                break
        else:
            self.root.after(0, lambda: self.image_label.config(image="", text="File not found"))
            return
        
        if self.current_file_path.lower().endswith(".xlsx"):
            for btn in [self.front_button, self.back_button, self.front_button_exp, self.back_button_exp]:
                btn.config(state="normal")
            self.on_page_click("Front")
        else:
            for btn in [self.front_button, self.back_button, self.front_button_exp, self.back_button_exp]:
                btn.config(state="disabled")
            self.display_file(self.current_file_path, "Image")
        
        self.root.after(2000, self.collapse_controls)
        if self.is_expanded:
            self.reset_collapse_timer()
    
    def display_file(self, path, page):
        if not path or path != self.current_file_path:
            return
        
        cache_key = f"{path}_{page}"
        cached = self.image_cache.get(cache_key)
        
        if cached:
            # Memory cache hit - instant display!
            if path == self.current_file_path:
                self.root.after(0, lambda: (self.image_label.config(image=cached, text=""),
                                           setattr(self.image_label, 'image', cached)))
        else:
            # Need to load from disk (may be in disk cache but still needs PIL processing)
            logger.info(f"Loading from disk: {os.path.basename(path)} - {page}")
            if path.lower().endswith(".xlsx"):
                self.root.after(0, lambda: self.image_label.config(image="", text=f"Loading {page}..."))
            else:
                self.root.after(0, lambda: self.image_label.config(image="", text="Loading..."))
            threading.Thread(target=self.load_file, args=(path, page, cache_key), daemon=True).start()
    
    def load_file(self, path, page, cache_key):
        """Load and display file - FIXED: proper return after error"""
        if path != self.current_file_path:
            logger.debug("Load cancelled - selection changed")
            return
        
        logger.info(f"Loading: {os.path.basename(path)} - {page}")
        
        try:
            if path.lower().endswith(".xlsx"):
                sheet_type = page.lower() if page != "Image" else "front"
                sheet = self.excel_converter.find_sheet(path, sheet_type)
                
                # CRITICAL FIX: Added proper return statement
                if not sheet:
                    logger.warning(f"Sheet type '{sheet_type}' not found in {os.path.basename(path)}")
                    if path == self.current_file_path:
                        self.root.after(0, lambda: self.image_label.config(image="", text="Sheet not found"))
                    return  # FIXED: This return was missing!
                
                png_path = self.excel_converter.convert_excel_to_png(path, sheet)
                
                if not png_path:
                    logger.error(f"Conversion failed for {os.path.basename(path)} - {sheet}")
                    if path == self.current_file_path:
                        self.root.after(0, lambda: self.image_label.config(image="", text=f"Failed to convert {page}"))
                    return
                
                img = Image.open(png_path)
            else:
                img = Image.open(path)
            
            screen_width = self.root.winfo_screenwidth()
            available_height = self.root.winfo_screenheight() - self.control_bar_collapsed_height
            
            max_dim = (min(screen_width, MAX_IMAGE_DIMENSION),
                      min(available_height, MAX_IMAGE_DIMENSION))
            img.thumbnail(max_dim, Image.LANCZOS)
            
            photo = ImageTk.PhotoImage(img)
            self.image_cache.put(cache_key, photo)
            logger.info(f"Stored in memory cache: {os.path.basename(path)} - {page}")
            
            if path == self.current_file_path:
                self.root.after(0, lambda: (self.image_label.config(image=photo, text=""),
                                           setattr(self.image_label, 'image', photo)))
        except Exception as e:
            logger.error(f"Error loading file {os.path.basename(path)}: {e}")
            if path == self.current_file_path:
                self.root.after(0, lambda: self.image_label.config(image="", text=f"Error loading:\n{os.path.basename(path)}"))
    
    def precache_dept(self, dept, stop_event):
        """Background precaching for entire department"""
        logger.info(f"=== BG PRECACHE START: {dept} ===")
        
        # Check if this is a special department with a custom path
        if dept in SPECIAL_DEPT_PATHS:
            dept_path = os.path.join(NETWORK_BASE_PATH, SPECIAL_DEPT_PATHS[dept])
        else:
            dept_path = os.path.join(NETWORK_BASE_PATH, dept)
        
        if not os.path.exists(dept_path):
            logger.warning(f"Dept path not found: {dept_path}")
            return
        
        try:
            # Get all model directories
            all_models = [d for d in os.listdir(dept_path)
                         if os.path.isdir(os.path.join(dept_path, d))]
            
            # Filter out excluded folders for this department
            excluded = EXCLUDED_MODEL_FOLDERS.get(dept, [])
            models = [m for m in all_models if m not in excluded]
            
            logger.info(f"BG precache: {len(models)} models in {dept}")
            
            for idx, model in enumerate(models, 1):
                if stop_event.is_set():
                    logger.info(f"BG precache cancelled at model {idx}/{len(models)}")
                    return
                
                model_path = os.path.join(dept_path, model)
                
                try:
                    files = [os.path.join(model_path, f)
                            for f in os.listdir(model_path)
                            if f.lower().endswith(".xlsx")]
                    
                    for excel_file in files:
                        if stop_event.is_set():
                            return
                        
                        for sheet_type in ["front", "back"]:
                            if stop_event.is_set():
                                return
                            
                            sheet = self.excel_converter.find_sheet(excel_file, sheet_type)
                            if sheet:
                                self.excel_converter.convert_excel_to_png(excel_file, sheet, stop_event)
                except Exception as e:
                    logger.debug(f"BG precache error in {model}: {e}")
                    continue
        except Exception as e:
            logger.error(f"BG precache error: {e}")
        
        logger.info(f"=== BG PRECACHE COMPLETE: {dept} ===")
    
    def precache_model_aggressive(self, model_path, stop_event):
        """Foreground precaching for current model"""
        model_name = os.path.basename(model_path)
        logger.info(f"=== FG PRECACHE START: {model_name} ===")
        
        if not os.path.exists(model_path):
            logger.warning(f"Model path not found: {model_path}")
            return
        
        try:
            files = [os.path.join(model_path, f)
                    for f in os.listdir(model_path)
                    if f.lower().endswith(".xlsx")]
            
            logger.info(f"FG precache: {len(files)} Excel files in {model_name}")
            
            for idx, excel_file in enumerate(files, 1):
                if stop_event.is_set():
                    logger.info(f"FG precache cancelled at file {idx}/{len(files)}")
                    return
                
                for sheet_type in ["front", "back"]:
                    if stop_event.is_set():
                        return
                    
                    sheet = self.excel_converter.find_sheet(excel_file, sheet_type)
                    if sheet:
                        self.excel_converter.convert_excel_to_png(excel_file, sheet, stop_event)
        except Exception as e:
            logger.error(f"FG precache error: {e}")
        
        logger.info(f"=== FG PRECACHE COMPLETE: {model_name} ===")

if __name__ == "__main__":
    root = tk.Tk()
    app = FullscreenImageApp(root)
    root.mainloop()
