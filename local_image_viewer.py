import tkinter as tk
from tkinter import ttk
import os
from PIL import Image, ImageTk
import threading
import time
from collections import Counter
import subprocess
import logging
from logging.handlers import RotatingFileHandler
from datetime import datetime, timedelta
import tempfile
import shutil
import gc

LOG_FILE = "/var/log/pi-photo-viewer/app.log"
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

try:
    handler = RotatingFileHandler(LOG_FILE, maxBytes=5*1024*1024, backupCount=5)
    formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
    handler.setFormatter(formatter)
    logger.addHandler(handler)
except:
    pass

USB_BASE_PATH = "/media/pi"
SUPPORTED_FORMATS = (".xlsx", ".png", ".jpg", ".jpeg", ".gif", ".bmp")
FOLDERS_TO_IGNORE = ["system volume information", "$recycle.bin"]
LOGO_WIDTH = 175
FALLBACK_SCAN_INTERVAL = 30
IMAGE_CACHE_SIZE = 2
MAX_IMAGE_DIMENSION = 1920
CACHE_STALE_DAYS = 7

SHEET_MAPPING = {
    "front": ["front", "front page", "proposal"],
    "back": ["back", "back page"],
    "hidden": ["changelog", "revision history"]
}

class ImageCache:
    def __init__(self, max_size=IMAGE_CACHE_SIZE):
        self.cache = {}
        self.max_size = max_size
        self.order = []
    
    def get(self, path):
        if path in self.cache:
            self.order.remove(path)
            self.order.append(path)
            return self.cache[path]
        return None
    
    def put(self, path, photo):
        if path in self.cache:
            self.order.remove(path)
        elif len(self.cache) >= self.max_size:
            removed = self.order.pop(0)
            if removed in self.cache:
                del self.cache[removed]
            gc.collect()
        self.cache[path] = photo
        self.order.append(path)
    
    def clear(self):
        self.cache.clear()
        self.order.clear()
        gc.collect()

class ExcelConverter:
    def __init__(self, cache_dir="/tmp/pi-photo-viewer-cache"):
        self.cache_dir = cache_dir
        self.conversion_lock = threading.Lock()  # Prevent concurrent conversions
        os.makedirs(cache_dir, exist_ok=True)
    
    def find_sheet(self, excel_path, sheet_type):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            patterns = SHEET_MAPPING.get(sheet_type.lower(), [])
            for pattern in patterns:
                for sheet_name in sheet_names:
                    if pattern.lower() in sheet_name.lower():
                        logger.info(f"Matched '{sheet_type}' to '{sheet_name}'")
                        return sheet_name
            return None
        except Exception as e:
            logger.error(f"Error reading sheet names: {e}")
            return None
    
    def is_cache_valid(self, cache_path, source_excel_path=None):
        """Check if cache is valid by comparing:
        1. Source Excel file modification time (auto-refresh if modified)
        2. Cache file age (fallback refresh after 7 days)
        """
        if not os.path.exists(cache_path):
            return False
        
        # Check if source Excel was modified since cache was created
        if source_excel_path and os.path.exists(source_excel_path):
            try:
                # Get source Excel modification time
                excel_mod_time = os.path.getmtime(source_excel_path)
                
                # Get cached metadata (if it exists)
                meta_path = cache_path.replace(".png", ".meta")
                if os.path.exists(meta_path):
                    with open(meta_path, 'r') as f:
                        cached_excel_mod_time = float(f.read().strip())
                    
                    # If Excel was modified after cache was created, cache is stale
                    if excel_mod_time > cached_excel_mod_time:
                        logger.info(f"Excel file modified, cache is stale: {source_excel_path}")
                        return False
            except Exception as e:
                logger.error(f"Error checking cache validity: {e}")
        
        # Check if cache PNG is older than 7 days (fallback refresh)
        try:
            file_age = datetime.now() - datetime.fromtimestamp(os.path.getmtime(cache_path))
            is_stale = file_age >= timedelta(days=CACHE_STALE_DAYS)
            if is_stale:
                logger.info(f"Cache file older than {CACHE_STALE_DAYS} days, refreshing")
            return not is_stale
        except Exception as e:
            logger.error(f"Error checking cache age: {e}")
            return False
    
    def save_cache_metadata(self, cache_path, source_excel_path):
        """Save source Excel mod time to metadata file"""
        try:
            excel_mod_time = os.path.getmtime(source_excel_path)
            meta_path = cache_path.replace(".png", ".meta")
            with open(meta_path, 'w') as f:
                f.write(str(excel_mod_time))
            logger.info(f"Saved cache metadata for {cache_path}")
        except Exception as e:
            logger.error(f"Error saving cache metadata: {e}")
    
    def get_cache_path(self, excel_path, sheet_name):
        safe_name = f"{os.path.basename(excel_path)}_{sheet_name}".replace(" ", "_")
        return os.path.join(self.cache_dir, f"{safe_name}.png")
    
    def get_cache_metadata_path(self, cache_png_path):
        """Get path to metadata file that tracks source Excel mod time"""
        return cache_png_path.replace(".png", ".meta")
    
    def get_sheet_index(self, excel_path, sheet_name):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            if sheet_name in sheet_names:
                return sheet_names.index(sheet_name)
            return None
        except Exception as e:
            logger.error(f"Error getting sheet index: {e}")
            return None
    
    def convert_excel_to_png(self, excel_path, sheet_name):
        temp_dir = None
        try:
            cache_path = self.get_cache_path(excel_path, sheet_name)
            if self.is_cache_valid(cache_path):
                logger.info(f"Using cached image: {cache_path}")
                return cache_path
            
            logger.info(f"Converting Excel sheet '{sheet_name}' to PNG...")
            sheet_index = self.get_sheet_index(excel_path, sheet_name)
            if sheet_index is None:
                return None
            
            temp_dir = tempfile.mkdtemp()
            logger.info(f"Created temp dir: {temp_dir}")
            
            try:
                output_prefix = cache_path.replace(".png", "")
                cmd = ["libreoffice", "--headless", "--invisible", "--nocrashreport", 
                       "--nodefault", "--nofirststartwizard", "--nologo", "--norestore",
                       "--convert-to", "pdf", "--outdir", temp_dir, excel_path]
                
                logger.info("Starting LibreOffice conversion...")
                result = subprocess.run(cmd, capture_output=True, timeout=45, text=True)
                
                if result.returncode != 0:
                    logger.error(f"LibreOffice conversion failed: {result.stderr}")
                    return None
                
                pdf_files = [f for f in os.listdir(temp_dir) if f.endswith(".pdf")]
                if not pdf_files:
                    logger.error("No PDF generated")
                    return None
                
                pdf_path = os.path.join(temp_dir, pdf_files[0])
                pdf_page = sheet_index + 1
                logger.info(f"PDF created at {pdf_path}, extracting page {pdf_page}...")
                
                cmd = ["pdftoppm", "-png", "-f", str(pdf_page), "-l", str(pdf_page), 
                       "-singlefile", "-r", "150", pdf_path, output_prefix]
                result = subprocess.run(cmd, capture_output=True, timeout=30, text=True)
                
                if result.returncode != 0:
                    logger.warning(f"pdftoppm failed: {result.stderr}")
                    logger.info("Trying ImageMagick fallback...")
                    cmd = ["convert", "-density", "100", f"{pdf_path}[{sheet_index}]", cache_path]
                    result = subprocess.run(cmd, capture_output=True, timeout=30, text=True)
                    if result.returncode != 0:
                        logger.error(f"ImageMagick also failed: {result.stderr}")
                        return None
                
                if os.path.exists(cache_path):
                    logger.info(f"Successfully converted to PNG: {cache_path}")
                    self.save_cache_metadata(cache_path, excel_path)
                    return cache_path
                else:
                    logger.error("PNG file was not created")
                    return None
            except subprocess.TimeoutExpired:
                logger.error("Conversion timed out")
                return None
            except Exception as e:
                logger.error(f"Conversion step error: {e}")
                return None
        except Exception as e:
            logger.error(f"Conversion error: {e}")
            return None
        finally:
            if temp_dir and os.path.exists(temp_dir):
                try:
                    logger.info(f"Cleaning up temp dir: {temp_dir}")
                    shutil.rmtree(temp_dir, ignore_errors=True)
                    time.sleep(0.1)  # Give OS time to release resources
                except Exception as e:
                    logger.error(f"Failed to clean temp dir: {e}")
            gc.collect()

class MediaWatcher:
    def __init__(self, callback, base_path=USB_BASE_PATH):
        self.callback = callback
        self.base_path = base_path
        self.last_state = None
        self.use_inotify = self._check_inotify()
        self.stop_event = threading.Event()
    
    def _check_inotify(self):
        try:
            subprocess.run(["which", "inotifywait"], capture_output=True, timeout=2, check=True)
            logger.info("inotifywait found - using real-time USB detection")
            return True
        except:
            logger.warning("inotifywait not found - falling back to polling")
            return False
    
    def get_media_state(self):
        state = {}
        if not os.path.exists(self.base_path):
            return state
        try:
            for drive in os.listdir(self.base_path):
                drive_path = os.path.join(self.base_path, drive)
                if os.path.isdir(drive_path):
                    try:
                        items = os.listdir(drive_path)
                        state[drive] = hash(tuple(sorted(items)))
                    except:
                        continue
        except:
            pass
        return state
    
    def watch_with_inotify(self):
        try:
            cmd = ["inotifywait", "-m", "-e", "create,delete,unmount", "-r", self.base_path]
            proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, bufsize=1)
            while not self.stop_event.is_set():
                try:
                    line = proc.stdout.readline()
                    if line:
                        logger.debug(f"inotify event: {line.strip()}")
                        self.callback()
                except Exception as e:
                    logger.error(f"inotify error: {e}")
                    break
            proc.terminate()
        except Exception as e:
            logger.error(f"Failed to use inotify: {e}")
            self.use_inotify = False
            self.watch_with_polling()
    
    def watch_with_polling(self):
        while not self.stop_event.is_set():
            current_state = self.get_media_state()
            if current_state != self.last_state:
                self.last_state = current_state
                logger.info("USB media state changed")
                self.callback()
            time.sleep(FALLBACK_SCAN_INTERVAL)
    
    def start(self):
        func = self.watch_with_inotify if self.use_inotify else self.watch_with_polling
        thread = threading.Thread(target=func, daemon=True)
        thread.start()
        return thread
    
    def stop(self):
        self.stop_event.set()

class FullscreenImageApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Pi Standards Viewer")
        logger.info("Application starting")
        
        self.root.after(200, lambda: self.root.attributes("-fullscreen", True))
        self.root.configure(bg="black")
        
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TCombobox", fieldbackground="#374151", background="#4B5563", 
                       foreground="#06B6D4", arrowcolor="#06B6D4", arrowsize=30)
        self.root.option_add("*TCombobox*Listbox.font", ("Helvetica", 22, "bold"))
        self.root.option_add("*TCombobox*Listbox.background", "#1F2937")
        self.root.option_add("*TCombobox*Listbox.foreground", "#06B6D4")
        
        self.control_bar_collapsed_height = 80
        self.control_bar_expanded_height = 180
        self.collapse_timer = None
        self.collapse_delay = 30000  # 30 seconds instead of 60
        
        self.control_frame = tk.Frame(root, bg="#1F2937", pady=10, padx=30)
        self.control_frame.pack(side="bottom", fill="x")
        self.control_frame.pack_propagate(False)
        self.control_frame.config(height=self.control_bar_collapsed_height)
        
        self.expand_indicator = tk.Label(self.control_frame, text="▲ TAP TO SELECT FILE ▲", 
                                         bg="#1F2937", fg="#06B6D4", font=("Helvetica", 14, "bold"), cursor="hand2")
        self.expand_indicator.pack(side="top", pady=(0, 5))
        self.expand_indicator.bind("<Button-1>", lambda e: self.expand_controls())
        
        self.collapsed_container = tk.Frame(self.control_frame, bg="#1F2937")
        self.collapsed_container.pack(fill="both", expand=True)
        
        self.expanded_container = tk.Frame(self.control_frame, bg="#1F2937")
        
        collapsed_label = tk.Label(self.collapsed_container, text="Page:", bg="#1F2937", 
                                   fg="#06B6D4", font=("Helvetica", 20, "bold"))
        collapsed_label.pack(side="left", padx=(0, 15))
        
        self.collapsed_button_frame = tk.Frame(self.collapsed_container, bg="#1F2937")
        self.collapsed_button_frame.pack(side="left", expand=True, fill="both")
        
        self.front_button = tk.Button(self.collapsed_button_frame, text="FRONT PAGE", 
                                     font=("Helvetica", 22, "bold"), bg="#06B6D4", fg="#1F2937",
                                     activebackground="#0891B2", relief="sunken", bd=3, cursor="hand2",
                                     command=lambda: self.on_page_button_click("Front"))
        self.back_button = tk.Button(self.collapsed_button_frame, text="BACK PAGE", 
                                    font=("Helvetica", 22, "bold"), bg="#4B5563", fg="#06B6D4",
                                    activebackground="#374151", relief="raised", bd=3, cursor="hand2",
                                    command=lambda: self.on_page_button_click("Back"))
        
        self.front_button.pack(side="left", expand=True, fill="both", padx=(0, 10))
        self.back_button.pack(side="left", expand=True, fill="both")
        
        exp_row1 = tk.Frame(self.expanded_container, bg="#1F2937")
        exp_row1.pack(fill="x", pady=(0, 10))
        tk.Label(exp_row1, text="Folder:", bg="#1F2937", fg="#06B6D4", 
                font=("Helvetica", 20, "bold")).pack(side="left", padx=(0, 10))
        self.folder_variable = tk.StringVar(root)
        self.folder_dropdown = ttk.Combobox(exp_row1, textvariable=self.folder_variable, 
                                           font=("Helvetica", 22, "bold"), state="readonly", 
                                           height=6, postcommand=lambda: self.on_dropdown_open())
        self.folder_dropdown.pack(side="left", expand=True, fill="both")
        self.folder_dropdown.bind("<<ComboboxSelected>>", self.on_folder_select)
        
        exp_row2 = tk.Frame(self.expanded_container, bg="#1F2937")
        exp_row2.pack(fill="x", pady=(0, 10))
        tk.Label(exp_row2, text="File:", bg="#1F2937", fg="#06B6D4", 
                font=("Helvetica", 20, "bold")).pack(side="left", padx=(0, 10))
        self.file_variable = tk.StringVar(root)
        self.file_dropdown = ttk.Combobox(exp_row2, textvariable=self.file_variable, 
                                         font=("Helvetica", 22, "bold"), state="readonly", 
                                         height=6, postcommand=lambda: self.on_dropdown_open())
        self.file_dropdown.pack(side="left", expand=True, fill="both")
        self.file_dropdown.bind("<<ComboboxSelected>>", self.on_file_select)
        
        exp_row3 = tk.Frame(self.expanded_container, bg="#1F2937")
        exp_row3.pack(fill="x")
        tk.Label(exp_row3, text="Page:", bg="#1F2937", fg="#06B6D4", 
                font=("Helvetica", 20, "bold")).pack(side="left", padx=(0, 10))
        self.expanded_button_frame = tk.Frame(exp_row3, bg="#1F2937")
        self.expanded_button_frame.pack(side="left", expand=True, fill="both")
        
        self.front_button_exp = tk.Button(self.expanded_button_frame, text="FRONT PAGE", 
                                         font=("Helvetica", 22, "bold"), bg="#06B6D4", fg="#1F2937",
                                         activebackground="#0891B2", relief="sunken", bd=3, cursor="hand2",
                                         command=lambda: self.on_page_button_click("Front"))
        self.back_button_exp = tk.Button(self.expanded_button_frame, text="BACK PAGE", 
                                        font=("Helvetica", 22, "bold"), bg="#4B5563", fg="#06B6D4",
                                        activebackground="#374151", relief="raised", bd=3, cursor="hand2",
                                        command=lambda: self.on_page_button_click("Back"))
        
        self.front_button_exp.pack(side="left", expand=True, fill="both", padx=(0, 10))
        self.back_button_exp.pack(side="left", expand=True, fill="both")
        
        self.image_label = tk.Label(root, bg="black", fg="white", font=("Helvetica", 24))
        self.image_label.pack(expand=True, fill="both")
        self.image_label.bind("<Button-1>", lambda e: self.expand_controls())
        
        self.logo_label = tk.Label(root, bg="black")
        self.logo_label.place(relx=1.0, rely=0.0, anchor="ne", x=-10, y=10)
        
        self.current_page = "Front"
        self.is_expanded = False
        self.display_to_path_map = {}
        self.files_by_folder = {}
        self.current_folder_path = None
        self.current_file_path = None
        self.image_cache = ImageCache()
        self.last_scan_result = None
        self.last_logo_path = None
        self.stop_threads = False
        self.excel_converter = ExcelConverter()
        self.precache_thread = None
        self.current_photoimage = None
        self.background_precache_thread = None
        self.stop_background_precache = False
        
        self.media_watcher = MediaWatcher(self.on_media_change)
        self.media_watcher.start()
        self.update_media_sources()
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
    
    def on_dropdown_open(self):
        if self.collapse_timer:
            self.root.after_cancel(self.collapse_timer)
            self.collapse_timer = None
        self.reset_collapse_timer()  # Reset timer when opening dropdown
    
    def precache_excel_files(self, folder_path):
        logger.info(f"Pre-caching Excel files in {folder_path}")
        try:
            files = self.files_by_folder.get(folder_path, [])
            excel_files = [f for f in files if f.lower().endswith(".xlsx")]
            for excel_file in excel_files:
                for sheet_type in ["front", "back"]:
                    actual_sheet = self.excel_converter.find_sheet(excel_file, sheet_type)
                    if actual_sheet:
                        logger.info(f"Pre-caching {os.path.basename(excel_file)} - {sheet_type}")
                        self.excel_converter.convert_excel_to_png(excel_file, actual_sheet)
            logger.info("Pre-caching complete")
        except Exception as e:
            logger.error(f"Pre-cache error: {e}")
    
    def background_precache_all_folders(self):
        """Slowly precache all folders in background"""
        logger.info("Starting background precache of all folders")
        try:
            # Precache all folders except current one (which is being done aggressively)
            for folder_path in sorted(self.files_by_folder.keys()):
                if self.stop_background_precache:
                    logger.info("Background precache stopped")
                    return
                
                logger.info(f"Background precaching folder: {folder_path}")
                try:
                    files = self.files_by_folder.get(folder_path, [])
                    excel_files = [f for f in files if f.lower().endswith(".xlsx")]
                    for excel_file in excel_files:
                        if self.stop_background_precache:
                            return
                        for sheet_type in ["front", "back"]:
                            if self.stop_background_precache:
                                return
                            actual_sheet = self.excel_converter.find_sheet(excel_file, sheet_type)
                            if actual_sheet:
                                logger.info(f"Background caching {os.path.basename(excel_file)} - {sheet_type}")
                                self.excel_converter.convert_excel_to_png(excel_file, actual_sheet)
                            time.sleep(0.5)  # Small delay between conversions
                except Exception as e:
                    logger.error(f"Background precache folder error: {e}")
                    continue
            logger.info("Background precache of all folders complete")
        except Exception as e:
            logger.error(f"Background precache error: {e}")
    
    def expand_controls(self):
        if self.is_expanded:
            return
        self.is_expanded = True
        self.collapsed_container.pack_forget()
        self.expanded_container.pack(fill="both", expand=True)
        self.expand_indicator.config(text="▼ TAP IMAGE TO HIDE ▼")
        self.control_frame.config(height=self.control_bar_expanded_height)
        self.reset_collapse_timer()
    
    def collapse_controls(self):
        if not self.is_expanded:
            return
        self.is_expanded = False
        self.expanded_container.pack_forget()
        self.collapsed_container.pack(fill="both", expand=True)
        self.expand_indicator.config(text="▲ TAP TO SELECT FILE ▲")
        self.control_frame.config(height=self.control_bar_collapsed_height)
        if self.collapse_timer:
            self.root.after_cancel(self.collapse_timer)
            self.collapse_timer = None
    
    def reset_collapse_timer(self):
        if self.collapse_timer:
            self.root.after_cancel(self.collapse_timer)
        self.collapse_timer = self.root.after(self.collapse_delay, self.collapse_controls)
    
    def on_page_button_click(self, page):
        self.current_page = page
        buttons = [(self.front_button, self.back_button), (self.front_button_exp, self.back_button_exp)]
        for front_btn, back_btn in buttons:
            if page == "Front":
                front_btn.config(bg="#06B6D4", fg="#1F2937", relief="sunken")
                back_btn.config(bg="#4B5563", fg="#06B6D4", relief="raised")
            else:
                back_btn.config(bg="#06B6D4", fg="#1F2937", relief="sunken")
                front_btn.config(bg="#4B5563", fg="#06B6D4", relief="raised")
        
        if self.current_file_path:
            self.display_file_async(self.current_file_path, page)
        if self.is_expanded:
            self.reset_collapse_timer()  # Reset timer on button click
    
    def on_media_change(self):
        self.root.after(500, self.update_media_sources)
    
    def on_closing(self):
        logger.info("Application closing")
        self.stop_threads = True
        self.stop_background_precache = True
        self.media_watcher.stop()
        self.root.destroy()
    
    def update_media_sources(self):
        new_display_to_path, new_files_by_folder, found_logo_path = self.scan_media_folders()
        new_scan_result = (tuple(sorted(new_display_to_path.items())), 
                          tuple(sorted((k, tuple(v)) for k, v in new_files_by_folder.items())),
                          found_logo_path)
        
        if new_scan_result == self.last_scan_result:
            return
        
        self.last_scan_result = new_scan_result
        logger.info(f"Media sources updated: {len(new_display_to_path)} folders found")
        self.display_to_path_map = new_display_to_path
        self.files_by_folder = new_files_by_folder
        
        # Start background precaching of all folders
        if self.background_precache_thread is None or not self.background_precache_thread.is_alive():
            self.stop_background_precache = False
            self.background_precache_thread = threading.Thread(
                target=self.background_precache_all_folders, 
                daemon=True
            )
            self.background_precache_thread.start()
        
        if found_logo_path != self.last_logo_path:
            self.last_logo_path = found_logo_path
            if found_logo_path:
                logger.info(f"Logo found: {found_logo_path}")
                logo_thread = threading.Thread(target=self.load_logo_threaded, args=(found_logo_path,), daemon=True)
                logo_thread.start()
            else:
                self.logo_label.config(image="")
                self.logo_label.image = None
        
        last_selected_folder = self.folder_variable.get()
        display_names = sorted(list(self.display_to_path_map.keys()))
        self.folder_dropdown["values"] = display_names
        
        if last_selected_folder in display_names:
            self.folder_variable.set(last_selected_folder)
        elif display_names:
            self.folder_variable.set(display_names[0])
        else:
            self.folder_variable.set("")
        self.update_file_list()
    
    def scan_media_folders(self):
        potential_folders = []
        found_logo_path = None
        new_display_to_path = {}
        new_files_by_folder = {}
        
        if os.path.exists(USB_BASE_PATH):
            try:
                drive_names = os.listdir(USB_BASE_PATH)
            except OSError:
                return new_display_to_path, new_files_by_folder, found_logo_path
            
            for drive_name in drive_names:
                drive_path = os.path.join(USB_BASE_PATH, drive_name)
                if not os.path.isdir(drive_path):
                    continue
                
                try:
                    if not found_logo_path:
                        try:
                            items = os.listdir(drive_path)
                            for item in items:
                                if item.lower() == "logo.png":
                                    found_logo_path = os.path.join(drive_path, item)
                                    break
                        except OSError:
                            continue
                    
                    try:
                        folder_names = os.listdir(drive_path)
                    except OSError:
                        continue
                    
                    for folder_name in folder_names:
                        if folder_name.lower() in FOLDERS_TO_IGNORE:
                            continue
                        folder_path = os.path.join(drive_path, folder_name)
                        if os.path.isdir(folder_path):
                            potential_folders.append({"drive": drive_name, "folder": folder_name, "path": folder_path})
                except (PermissionError, OSError):
                    continue
        
        folder_counts = Counter(item["folder"] for item in potential_folders)
        for item in potential_folders:
            folder_name = item["folder"]
            display_name = folder_name
            if folder_counts[folder_name] > 1:
                display_name = "{} ({})".format(folder_name, item["drive"])
            
            folder_path = item["path"]
            new_display_to_path[display_name] = folder_path
            new_files_by_folder[folder_path] = []
            
            try:
                for file_item in os.listdir(folder_path):
                    full_item_path = os.path.join(folder_path, file_item)
                    try:
                        if os.path.isfile(full_item_path) and file_item.lower().endswith(SUPPORTED_FORMATS):
                            new_files_by_folder[folder_path].append(full_item_path)
                    except OSError:
                        continue
                new_files_by_folder[folder_path].sort()
            except (PermissionError, OSError):
                continue
        
        return new_display_to_path, new_files_by_folder, found_logo_path
    
    def on_folder_select(self, event):
        self.update_file_list()
        if self.is_expanded:
            self.reset_collapse_timer()  # Reset timer on folder selection
    
    def update_file_list(self):
        selected_folder_display_name = self.folder_variable.get()
        if not selected_folder_display_name:
            self.file_dropdown["values"] = []
            self.file_variable.set("")
            self.image_label.config(image="", text="No media found.")
            return
        
        folder_path = self.display_to_path_map.get(selected_folder_display_name)
        current_files = self.files_by_folder.get(folder_path, [])
        self.current_folder_path = folder_path
        
        if current_files:
            display_filenames = [os.path.splitext(os.path.basename(p))[0] for p in current_files]
            self.file_dropdown["values"] = display_filenames
            self.file_variable.set(display_filenames[0])
            self.on_file_select(None)
            
            if self.precache_thread is None or not self.precache_thread.is_alive():
                self.precache_thread = threading.Thread(target=self.precache_excel_files, args=(folder_path,), daemon=True)
                self.precache_thread.start()
        else:
            self.file_dropdown["values"] = []
            self.file_variable.set("")
            self.image_label.config(image="", text=f"No files found in\n{selected_folder_display_name}")
    
    def on_file_select(self, event):
        selected_file_name = self.file_variable.get()
        if not selected_file_name or not self.current_folder_path:
            return
        
        current_files = self.files_by_folder.get(self.current_folder_path, [])
        file_path = None
        for f in current_files:
            if os.path.splitext(os.path.basename(f))[0] == selected_file_name:
                file_path = f
                break
        
        if not file_path:
            return
        
        self.current_file_path = file_path
        
        if file_path.lower().endswith(".xlsx"):
            for btn in [self.front_button, self.back_button, self.front_button_exp, self.back_button_exp]:
                btn.config(state="normal")
            self.on_page_button_click("Front")
        else:
            for btn in [self.front_button, self.back_button, self.front_button_exp, self.back_button_exp]:
                btn.config(state="disabled")
            self.display_file_async(file_path, "Image")
        
        self.root.after(2000, self.collapse_controls)
        if self.is_expanded:
            self.reset_collapse_timer()
    
    def display_file_async(self, file_path, page):
        if not file_path:
            return
        
        cache_key = f"{file_path}_{page}"
        cached = self.image_cache.get(cache_key)
        
        if cached:
            self.update_ui_with_image(cached)
        else:
            if file_path.lower().endswith(".xlsx"):
                self.image_label.config(image="", text=f"Converting Excel sheet: {page}\n\nThis may take 3-5 seconds on first load...\nSubsequent loads will be instant.")
            else:
                self.image_label.config(image="", text="Loading...")
            thread = threading.Thread(target=self.load_file_threaded, args=(file_path, page, cache_key), daemon=True)
            thread.start()
    
    def load_file_threaded(self, file_path, page, cache_key):
        try:
            logger.info(f"load_file_threaded START: {file_path}, page={page}")
            if file_path.lower().endswith(".xlsx"):
                logger.info("Processing Excel file")
                sheet_type = page.lower() if page != "Image" else "front"
                logger.info(f"Sheet type: {sheet_type}")
                actual_sheet = self.excel_converter.find_sheet(file_path, sheet_type)
                
                if not actual_sheet:
                    logger.error(f"Could not find sheet: {sheet_type}")
                    self.root.after(0, self.update_ui_with_error, f"Could not find '{page}' sheet")
                    return
                
                logger.info(f"Found sheet: {actual_sheet}")
                png_path = self.excel_converter.convert_excel_to_png(file_path, actual_sheet)
                logger.info(f"Conversion result: {png_path}")
                
                if not png_path:
                    logger.error("Conversion returned None")
                    self.root.after(0, self.update_ui_with_error, f"Failed to convert {page}")
                    return
                
                try:
                    logger.info(f"Opening PNG: {png_path}")
                    img = Image.open(png_path)
                    logger.info(f"PNG opened successfully, size: {img.size}")
                except Exception as e:
                    logger.error(f"Failed to open PNG: {e}", exc_info=True)
                    self.root.after(0, self.update_ui_with_error, f"Error opening converted image")
                    return
            else:
                logger.info("Processing regular image file")
                try:
                    logger.info(f"Opening image: {file_path}")
                    img = Image.open(file_path)
                    logger.info(f"Image opened successfully, size: {img.size}")
                except Exception as e:
                    logger.error(f"Failed to open image: {e}", exc_info=True)
                    self.root.after(0, self.update_ui_with_error, os.path.basename(file_path))
                    return
            
            try:
                logger.info("Getting screen dimensions")
                screen_width = self.root.winfo_screenwidth()
                available_height = self.root.winfo_screenheight() - self.control_bar_collapsed_height
                logger.info(f"Screen: {screen_width}x{available_height}")
                
                max_dim = (min(screen_width, MAX_IMAGE_DIMENSION), 
                          min(available_height, MAX_IMAGE_DIMENSION))
                logger.info(f"Thumbnailing to: {max_dim}")
                img.thumbnail(max_dim, Image.LANCZOS)
                logger.info(f"Thumbnail created, new size: {img.size}")
                
                logger.info("Creating PhotoImage")
                photo = ImageTk.PhotoImage(img)
                logger.info("PhotoImage created successfully")
                
                logger.info("Adding to cache")
                self.image_cache.put(cache_key, photo)
                logger.info("Cache updated, scheduling UI update")
                
                self.root.after(0, self.update_ui_with_image, photo)
                logger.info("UI update scheduled")
            except Exception as e:
                logger.error(f"Error processing image: {e}", exc_info=True)
                self.root.after(0, self.update_ui_with_error, "Image processing error")
        except Exception as e:
            logger.error(f"Error loading file {file_path}: {e}", exc_info=True)
            self.root.after(0, self.update_ui_with_error, os.path.basename(file_path))
    
    def update_ui_with_image(self, photo):
        try:
            self.image_label.config(image=photo, text="")
            self.image_label.image = photo
        except Exception as e:
            logger.error(f"Error updating UI with image: {e}")
    
    def update_ui_with_error(self, filename):
        self.image_label.config(image="", text=f"Error loading:\n{filename}")
    
    def load_logo_threaded(self, path):
        try:
            img = Image.open(path)
            w, h = img.size
            new_height = int((LOGO_WIDTH / float(w)) * h)
            img = img.resize((LOGO_WIDTH, new_height), Image.LANCZOS)
            photo = ImageTk.PhotoImage(img)
            self.root.after(0, self.update_ui_with_logo, photo)
        except Exception as e:
            logger.error(f"Error loading logo: {e}")
    
    def update_ui_with_logo(self, photo):
        self.logo_label.config(image=photo)
        self.logo_label.image = photo

if __name__ == "__main__":
    root = tk.Tk()
    app = FullscreenImageApp(root)
    root.mainloop()
